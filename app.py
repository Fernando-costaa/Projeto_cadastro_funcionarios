from typing import Tuple
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib 
from openpyxl import Workbook


#colocar a aparencia padrao do sistema
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.tema()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de gestão de Funcionários")
        self.geometry("700x500")

    def tema(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50,corner_radius=0,bg_color="teal", fg_color="teal",).place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Cadastro de Funcionários", font=("Century Gothic bold", 24), text_color="#fff").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha todos os dados!", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=50, y=70)

        planilha = pathlib.Path("Funcionários.xlsx")
        
        if planilha.exists():
              pass
        else:
                planilha=Workbook()
                pagina=planilha.active
                pagina['A1']="Nome Completo"
                pagina['B1']="CPF"
                pagina['C1']="Data de Nascimento"
                pagina['D1']="Data de Admissão"
                pagina['E1']="Função"
                pagina['F1']="Cursos"

                planilha.save("Funcionários.xlsx")
        
        def submit():

            #pegando os dados das entradas
            nome = nome_value.get()
            cpf = cpf_value.get()
            funcao = funcao_value.get()
            data_admissao = dataadmissao_value.get()
            data_nascimento = datanascimento_value.get()
            aso = aso_combobox.get()

            planilha = openpyxl.load_workbook('Funcionários.xlsx')
            pagina = planilha.active
            pagina.cell(column=1, row=pagina.max_row+1, value=nome)
            pagina.cell(column=2, row=pagina.max_row, value=cpf)
            pagina.cell(column=3, row=pagina.max_row, value=data_nascimento)
            pagina.cell(column=4, row=pagina.max_row, value=data_admissao)
            pagina.cell(column=5, row=pagina.max_row, value=funcao)
            pagina.cell(column=6, row=pagina.max_row, value=aso)

            planilha.save(r"Funcionarios.xlsx")
            messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

           
        
        def clear():
            nome_value.set("")
            cpf_value.set("")
            funcao_value.set("")
            dataadmissao_value.set("")
            datanascimento_value.set("")


        #variaveis de texto
        nome_value = StringVar()
        cpf_value = StringVar()
        funcao_value = StringVar()
        dataadmissao_value = StringVar()
        datanascimento_value = StringVar()

        

        #Entradas
        nome_entry = ctk.CTkEntry(self, width=350, textvariable=nome_value, font=("Century Gothic bold", 16), fg_color="transparent")
        cpf_entry = ctk.CTkEntry(self, width=150, textvariable=cpf_value, font=("Century Gothic bold", 16), fg_color="transparent")
        funcao_entry = ctk.CTkEntry(self, width=200, textvariable=funcao_value, font=("Century Gothic bold", 16), fg_color="transparent")
        data_admissao_entry = ctk.CTkEntry(self, width=150, textvariable=dataadmissao_value, font=("Century Gothic bold", 16), fg_color="transparent")
        data_nascimento_entry = ctk.CTkEntry(self, width=140, textvariable=datanascimento_value, font=("Century Gothic bold", 16), fg_color="transparent")
       #combobox
        aso_combobox = ctk.CTkComboBox(self, values=["Nenhum", "12", "18", "12 e 18"], font=("Century Gothic bold", 14))
        aso_combobox.set("Nenhum")

        


        #labels
        lb_nome = ctk.CTkLabel(self, text="Nome Completo", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_cpf = ctk.CTkLabel(self, text="CPF", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_funcao = ctk.CTkLabel(self, text="Função", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_data_nascimento = ctk.CTkLabel(self, text="Data de Nascimento", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_data_admissao = ctk.CTkLabel(self, text="Data de Admissão", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_aso = ctk.CTkLabel(self, text="Cursos", font=("Century Gothic bold", 16), text_color=["#000","#fff"])

        btn_enviar = ctk.CTkButton(self, text= "Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_limpar = ctk.CTkButton(self, text= "Limpar Dados".upper(), command=clear, fg_color="#555", hover_color="#333")

        #posicionando os elementos
        lb_nome.place(x=50, y=120)
        nome_entry.place(x=50, y=150)

        lb_data_nascimento.place(x=450, y=120)
        data_nascimento_entry.place(x=450, y=150)

        lb_data_admissao.place(x=50, y=190)
        data_admissao_entry.place(x=50, y=220)

        lb_aso.place(x=450, y=190)
        aso_combobox.place(x=450, y=220)

        lb_funcao.place(x=230, y=190)
        funcao_entry.place(x=230, y=220)

        lb_cpf.place(x=50, y=260)
        cpf_entry.place(x=50, y=290)

        btn_enviar.place(x=200, y=420)
        btn_limpar.place(x=380, y=420)


        



    def change_apm(self, novo_tema):
        ctk.set_appearance_mode(novo_tema)

if __name__=="__main__":
    app = App()
    app.mainloop()